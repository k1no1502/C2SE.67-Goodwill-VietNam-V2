import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

# Tạo workbook mới
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sản Phẩm'

# Headers
headers = ['STT', 'Tên Vật Phẩm', 'Danh Mục', 'Mô Tả Chi Tiết', 'Số Lượng', 'Đơn Vị', 'Tình Trạng', 'Giá Trị Ước Tính (vnd)']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')

# Độ rộng cột
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 18

# Tạo sheet danh mục
dv_ws = wb.create_sheet('DanhMuc')
danhMuc = ['Điện tử', 'Đồ chơi', 'Gia dụng', 'Khác', 'Quần áo', 'Sách', 'Thực phẩm', 'Y tế']
for idx, cat in enumerate(danhMuc, 1):
    dv_ws.cell(row=idx, column=1).value = cat

# Tạo sheet đơn vị
dv_ws2 = wb.create_sheet('DonVi')
donVi = ['Cái', 'Chiếc', 'Bộ', 'Hộp', 'Bao', 'Đôi', 'Gói', 'Chai', 'Lọ', 'Bụng', 'Thùng', 'Khác']
for idx, unit in enumerate(donVi, 1):
    dv_ws2.cell(row=idx, column=1).value = unit

# Tạo sheet tình trạng
dv_ws3 = wb.create_sheet('TinhTrang')
tinhTrang = ['Tốt', 'Bình thường', 'Cần sửa', 'Hỏng']
for idx, status in enumerate(tinhTrang, 1):
    dv_ws3.cell(row=idx, column=1).value = status

# Ẩn các sheet
dv_ws.sheet_state = 'hidden'
dv_ws2.sheet_state = 'hidden'
dv_ws3.sheet_state = 'hidden'

# Data validation cho cột C (Danh Mục)
dv_cat = DataValidation(type='list', formula1="DanhMuc!$A$1:$A$8", allow_blank=False)
dv_cat.error = 'Vui lòng chọn danh mục từ danh sách'
dv_cat.errorTitle = 'Lỗi'
dv_cat.prompt = 'Chọn danh mục'
dv_cat.promptTitle = 'Danh Mục'
ws.add_data_validation(dv_cat)

# Data validation cho cột F (Đơn Vị)
dv_unit = DataValidation(type='list', formula1="DonVi!$A$1:$A$12", allow_blank=False)
dv_unit.error = 'Vui lòng chọn đơn vị từ danh sách'
dv_unit.errorTitle = 'Lỗi'
dv_unit.prompt = 'Chọn đơn vị'
dv_unit.promptTitle = 'Đơn Vị'
ws.add_data_validation(dv_unit)

# Data validation cho cột G (Tình Trạng)
dv_status = DataValidation(type='list', formula1="TinhTrang!$A$1:$A$4", allow_blank=False)
dv_status.error = 'Vui lòng chọn tình trạng từ danh sách'
dv_status.errorTitle = 'Lỗi'
dv_status.prompt = 'Chọn tình trạng'
dv_status.promptTitle = 'Tình Trạng'
ws.add_data_validation(dv_status)

# Áp dụng dropdown cho 100 dòng
for row in range(2, 101):
    dv_cat.add('C' + str(row))
    dv_unit.add('F' + str(row))
    dv_status.add('G' + str(row))

# Thêm example rows
ws.cell(row=2, column=1).value = 1

# Lưu file
output_path = 'SanPham_Template.xlsx'
wb.save(output_path)
print('✓ File Excel cập nhật thành công: SanPham_Template.xlsx')
print('  - 8 cột dữ liệu: STT, Tên, Danh Mục, Mô Tả, Số Lượng, Đơn Vị, Tình Trạng, Giá')
print('  - 3 dropdown: Danh Mục (8), Đơn Vị (12), Tình Trạng (4)')
print('  - 3 sheet ẩn: DanhMuc, DonVi, TinhTrang')

